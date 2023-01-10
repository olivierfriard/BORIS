"""
BORIS
Behavioral Observation Research Interactive Software
Copyright 2012-2020 Olivier Friard

module for testing export_observation.py

pytest -vv test_export_observation.py
"""

import pytest
import sys
import json
import os
from openpyxl import load_workbook
import tablib

sys.path.insert(0, os.path.abspath(os.path.join(os.path.dirname(__file__), "..")))

from boris import config
from boris import export_observation


SELECTED_SUBJECTS = config.SELECTED_SUBJECTS
SELECTED_BEHAVIORS = config.SELECTED_BEHAVIORS


@pytest.fixture()
def before():
    os.system("rm -rf output")
    os.system("mkdir output")


class Test_export_events(object):
    @pytest.mark.usefixtures("before")
    def test_export_tabular_tsv(self):

        pj = json.loads(open("files/test.boris").read())
        obs_id = "observation #1"
        parameters = {SELECTED_SUBJECTS: ["subject1", "subject2"], SELECTED_BEHAVIORS: ["p", "s"]}
        file_name = "test_export_events_tabular.tsv"
        output_format = "tsv"

        r, msg = export_observation.export_events(
            parameters,
            obs_id,
            pj[config.OBSERVATIONS][obs_id],
            pj[config.ETHOGRAM],
            "output/" + file_name,
            output_format,
        )

        assert (
            open("files/test_export_events_tabular.tsv").read() == open("output/test_export_events_tabular.tsv").read()
        )

    @pytest.mark.usefixtures("before")
    def test_export_tabular_csv(self):

        pj = json.loads(open("files/test.boris").read())
        obs_id = "observation #1"
        parameters = {SELECTED_SUBJECTS: ["subject1", "subject2"], SELECTED_BEHAVIORS: ["p", "s"]}
        file_name = "test_export_events_tabular.csv"
        output_format = "csv"

        r, msg = export_observation.export_events(
            parameters,
            obs_id,
            pj[config.OBSERVATIONS][obs_id],
            pj[config.ETHOGRAM],
            "output/" + file_name,
            output_format,
        )

        assert (
            open("files/test_export_events_tabular.csv").read() == open("output/test_export_events_tabular.csv").read()
        )

    @pytest.mark.usefixtures("before")
    def test_export_tabular_html(self):

        pj = json.loads(open("files/test.boris").read())
        obs_id = "observation #1"
        parameters = {SELECTED_SUBJECTS: ["subject1", "subject2"], SELECTED_BEHAVIORS: ["p", "s"]}
        file_name = "test_export_events_tabular.html"
        output_format = "html"

        r, msg = export_observation.export_events(
            parameters,
            obs_id,
            pj[config.OBSERVATIONS][obs_id],
            pj[config.ETHOGRAM],
            "output/" + file_name,
            output_format,
        )

        assert (
            open("files/test_export_events_tabular.html").read()
            == open("output/test_export_events_tabular.html").read()
        )

    @pytest.mark.usefixtures("before")
    def test_export_tabular_xlsx(self):

        pj = json.loads(open("files/test.boris").read())

        obs_id = "observation #1"
        parameters = {SELECTED_SUBJECTS: ["subject1", "subject2"], SELECTED_BEHAVIORS: ["p", "s"]}
        file_name = "test_export_events_tabular.xlsx"
        output_format = "xlsx"

        r, msg = export_observation.export_events(
            parameters,
            obs_id,
            pj[config.OBSERVATIONS][obs_id],
            pj[config.ETHOGRAM],
            "output/" + file_name,
            output_format,
        )

        ref_all_cells = []
        wb = load_workbook(filename=f"files/{file_name}", read_only=True)
        for ws_name in wb.sheetnames:
            ref_all_cells.extend([cell.value for row in wb[ws_name].rows for cell in row])

        test_all_cells = []
        wb = load_workbook(filename=f"output/{file_name}", read_only=True)
        for ws_name in wb.sheetnames:
            worksheet_name = ws_name
            test_all_cells.extend([cell.value for row in wb[ws_name].rows for cell in row])

        assert worksheet_name == obs_id
        assert ref_all_cells == test_all_cells


# TODO: add testing long worksheet title xls

# TODO: add testing for worksheet title containing forbidden characters xls/xlsx


class Test_export_aggregated_events(object):
    @pytest.mark.usefixtures("before")
    def test_full_1(self):
        """
        export aggregated events:
        no time interval restriction: 9 min
        all subjects
        all behaviors
        """
        pj = json.loads(open("files/test.boris").read())
        obs_id = "observation #2"
        parameters = {
            "selected subjects": ["subject1", "subject2", "No focal subject"],
            SELECTED_BEHAVIORS: ["p", "s"],
            "time": config.TIME_FULL_OBS,
            config.START_TIME: 0,
            config.END_TIME: 9 * 60,
        }

        tablib_dataset = export_observation.export_aggregated_events(pj, parameters, obs_id)
        tablib_dataset_tsv = tablib_dataset.tsv

        print(tablib_dataset_tsv)

        """
        print(tablib_dataset_tsv)
        open("files/test_export_aggregated_events_test_full_1.tsv","w").write(tablib_dataset_tsv)
        """

        ref = open("files/test_export_aggregated_events_test_full_1.tsv").read()
        assert tablib_dataset_tsv.replace("\r", "") == ref

    @pytest.mark.usefixtures("before")
    def test_full_2(self):
        """
        export aggregated events:
        no time interval restriction: 9 min
        1 subject / 2
        all behaviors
        """
        pj = json.loads(open("files/test.boris").read())
        obs_id = "observation #2"
        parameters = {
            "selected subjects": ["subject1"],
            "selected behaviors": ["p", "s"],
            "time": config.TIME_FULL_OBS,
            config.START_TIME: 0,
            config.END_TIME: 9 * 60,
        }

        tablib_dataset = export_observation.export_aggregated_events(pj, parameters, obs_id)
        tablib_dataset_tsv = tablib_dataset.tsv

        """
        print(tablib_dataset_json)
        open("files/test_export_aggregated_events_test_full_2.tsv","w").write(tablib_dataset_tsv)
        """

        ref = open("files/test_export_aggregated_events_test_full_2.tsv").read()
        assert tablib_dataset_tsv.replace("\r", "") == ref

    @pytest.mark.usefixtures("before")
    def test_full_3(self):
        """
        export aggregated events:
        no time interval restriction: 9 min
        all subjects
        1 behavior / 2
        """
        pj = json.loads(open("files/test.boris").read())
        obs_id = "observation #2"
        parameters = {
            "selected subjects": ["subject1", "subject2", "No focal subject"],
            "selected behaviors": ["s"],
            "time": config.TIME_FULL_OBS,
            config.START_TIME: 0,
            config.END_TIME: 9 * 60,
        }

        tablib_dataset = export_observation.export_aggregated_events(pj, parameters, obs_id)
        tablib_dataset_tsv = tablib_dataset.tsv

        """
        print(tablib_dataset_tsv)
        open("files/test_export_aggregated_events_test_full_3.tsv","w").write(tablib_dataset_tsv)
        """

        ref = open("files/test_export_aggregated_events_test_full_3.tsv").read()
        assert tablib_dataset_tsv.replace("\r", "") == ref

    @pytest.mark.usefixtures("before")
    def test_full_with_trailing_spaces_in_modifiers(self):
        """
        some modifiers were configured with trailing spaces
        """
        pj = json.loads(open("files/test_with_leading_trailing_spaces_in_modifiers.boris").read())
        obs_id = "test1 live"
        parameters = {
            "selected subjects": ["No focal subject"],
            "selected behaviors": ["p", "r"],
            "time": config.TIME_FULL_OBS,
            config.START_TIME: 0,
            config.END_TIME: 2 * 60,
        }

        tablib_dataset = export_observation.export_aggregated_events(pj, parameters, obs_id)
        tablib_dataset_tsv = tablib_dataset.tsv

        """
        print(tablib_dataset_tsv)
        open("files/test_export_aggregated_events_test_full_with_trailing_spaces_in_modifiers.tsv", "w").write(tablib_dataset_tsv)
        """

        ref = open("files/test_export_aggregated_events_test_full_with_trailing_spaces_in_modifiers.tsv").read()
        assert tablib_dataset_tsv.replace("\r", "") == ref

    @pytest.mark.usefixtures("before")
    def test_partial_4(self):
        """
        export aggregated events:
        time interval restriction: 0 - 60 s
        all subjects
        all behaviors
        """
        pj = json.loads(open("files/test.boris").read())
        obs_id = "observation #2"
        parameters = {
            "selected subjects": ["subject1", "subject2", "No focal subject"],
            "selected behaviors": ["s", "p"],
            "time": config.TIME_ARBITRARY_INTERVAL,
            config.START_TIME: 0,
            config.END_TIME: 1 * 60,
        }

        tablib_dataset = export_observation.export_aggregated_events(pj, parameters, obs_id)
        tablib_dataset_tsv = tablib_dataset.tsv

        """
        print(tablib_dataset_tsv)
        open("files/test_export_aggregated_events_test_full_4.tsv","w").write(tablib_dataset_tsv)
        """

        ref = open("files/test_export_aggregated_events_test_full_4.tsv").read()
        assert tablib_dataset_tsv.replace("\r", "") == ref

    @pytest.mark.usefixtures("before")
    def test_partial_5(self):
        """
        export aggregated events:
        time interval restriction: 5 - 30 s
        all subjects
        all behaviors
        """
        pj = json.loads(open("files/test.boris").read())
        obs_id = "observation #2"
        parameters = {
            "selected subjects": ["subject1", "subject2", "No focal subject"],
            "selected behaviors": ["s", "p"],
            "time": config.TIME_ARBITRARY_INTERVAL,
            config.START_TIME: 5,
            config.END_TIME: 30,
        }

        tablib_dataset = export_observation.export_aggregated_events(pj, parameters, obs_id)
        tablib_dataset_tsv = tablib_dataset.tsv

        print(tablib_dataset_tsv)
        open("files/test_export_aggregated_events_test_full_5.tsv", "w").write(tablib_dataset_tsv)

        ref = open("files/test_export_aggregated_events_test_full_5.tsv").read()
        assert tablib_dataset_tsv.replace("\r", "") == ref

    @pytest.mark.usefixtures("before")
    def test_partial_6(self):
        """
        export aggregated events: No events
        time interval restriction: 60 - 180 s
        all subjects
        all behaviors
        """
        pj = json.loads(open("files/test.boris").read())
        obs_id = "observation #2"
        parameters = {
            config.SELECTED_SUBJECTS: ["subject1", "subject2", "No focal subject"],
            config.SELECTED_BEHAVIORS: ["s", "p"],
            "time": config.TIME_ARBITRARY_INTERVAL,
            config.START_TIME: 60,
            config.END_TIME: 180,
        }

        tablib_dataset = export_observation.export_aggregated_events(pj, parameters, obs_id)
        tablib_dataset_tsv = tablib_dataset.tsv

        """
        print(tablib_dataset_tsv)
        open("files/test_export_aggregated_events_test_full_6.tsv","w").write(tablib_dataset_tsv)
        """

        ref = open("files/test_export_aggregated_events_test_full_6.tsv").read()
        assert tablib_dataset_tsv.replace("\r", "") == ref


class Test_export_events_jwatcher(object):
    @pytest.mark.usefixtures("before")
    def test_1(self):

        pj = json.loads(open("files/test.boris").read())

        obs_id = "observation #1"
        parameters = {config.SELECTED_SUBJECTS: ["subject1"], config.SELECTED_BEHAVIORS: ["p", "s"]}
        file_name = "test_jwatcher"
        output_format = ""

        r, msg = export_observation.export_events_jwatcher(
            parameters, obs_id, pj["observations"][obs_id], pj[config.ETHOGRAM], "output/" + file_name, output_format
        )

        ref = [x for x in open("files/test_jwatcher_subject1.dat").readlines() if not x.startswith("#")]
        out = [x for x in open("output/test_jwatcher_subject1.dat").readlines() if not x.startswith("#")]
        assert ref == out

        ref = [x for x in open("files/test_jwatcher_subject1.faf").readlines() if not x.startswith("#")]
        out = [x for x in open("output/test_jwatcher_subject1.faf").readlines() if not x.startswith("#")]
        assert ref == out

        ref = [x for x in open("files/test_jwatcher_subject1.fmf").readlines() if not x.startswith("#")]
        out = [x for x in open("output/test_jwatcher_subject1.fmf").readlines() if not x.startswith("#")]
        assert ref == out


class Test_events_to_behavioral_sequences(object):
    def test_1(self):

        pj = json.loads(open("files/test.boris").read())

        obs_id = "observation #1"
        subject = "subject1"
        parameters = {
            SELECTED_SUBJECTS: ["subject1"],
            SELECTED_BEHAVIORS: ["p", "s"],
            config.INCLUDE_MODIFIERS: False,
            config.EXCLUDE_BEHAVIORS: False,
            config.START_TIME: 0,
            config.END_TIME: 100.0,
        }

        behav_seq_separator = "|"

        out = export_observation.events_to_behavioral_sequences(pj, obs_id, subject, parameters, behav_seq_separator)

        assert open("files/Test_events_to_behavioral_sequences_test_1").read() == out

    def test_2_separator_changed(self):

        pj = json.loads(open("files/test.boris").read())

        obs_id = "observation #1"
        subject = "subject1"
        parameters = {
            SELECTED_SUBJECTS: ["subject1"],
            SELECTED_BEHAVIORS: ["p", "s"],
            config.INCLUDE_MODIFIERS: False,
            config.EXCLUDE_BEHAVIORS: False,
            config.START_TIME: 0,
            config.END_TIME: 100.0,
        }

        behav_seq_separator = "£"

        out = export_observation.events_to_behavioral_sequences(pj, obs_id, subject, parameters, behav_seq_separator)

        assert open("files/Test_events_to_behavioral_sequences_test_2_separator").read() == out

    def test_3_no_behavior_found_for_selected_subject(self):

        pj = json.loads(open("files/test.boris").read())

        obs_id = "observation #1"
        subject = "subject1"
        parameters = {
            SELECTED_SUBJECTS: ["subject2"],
            SELECTED_BEHAVIORS: ["p"],
            config.INCLUDE_MODIFIERS: False,
            config.EXCLUDE_BEHAVIORS: False,
            config.START_TIME: 0,
            config.END_TIME: 100.0,
        }

        behav_seq_separator = "|"

        out = export_observation.events_to_behavioral_sequences(pj, obs_id, subject, parameters, behav_seq_separator)

        # open("1", "w").write(out)
        assert out == ""

    def test_4_behaviors_with_modifiers(self):

        pj = json.loads(open("files/test.boris").read())

        obs_id = "modifiers"
        subject = ""
        parameters = {
            SELECTED_SUBJECTS: [""],
            SELECTED_BEHAVIORS: ["q", "r"],
            config.INCLUDE_MODIFIERS: True,
            config.EXCLUDE_BEHAVIORS: False,
            config.START_TIME: 0,
            config.END_TIME: 100.0,
        }

        behav_seq_separator = "|"

        out = export_observation.events_to_behavioral_sequences(pj, obs_id, subject, parameters, behav_seq_separator)

        assert open("files/Test_events_to_behavioral_sequences_test_4_behaviors_with_modifiers").read() == out

    def test_5_observation_not_paired(self):

        pj = json.loads(open("files/test.boris").read())

        obs_id = "live not paired"
        subject = ""
        parameters = {
            SELECTED_SUBJECTS: [""],
            SELECTED_BEHAVIORS: ["p", "s"],
            config.INCLUDE_MODIFIERS: False,
            config.EXCLUDE_BEHAVIORS: False,
            config.START_TIME: 0,
            config.END_TIME: 100.0,
        }

        behav_seq_separator = "|"

        out = export_observation.events_to_behavioral_sequences(pj, obs_id, subject, parameters, behav_seq_separator)

        # open("1", "w").write(out)
        assert open("files/Test_events_to_behavioral_sequences_test_5_observation_not_paired").read() == out

    @pytest.mark.usefixtures("before")
    def test_6_multirow_description(self):

        pj = json.loads(open("files/test.boris").read())

        observations = ["live export behavioral sequences"]
        parameters = {
            SELECTED_SUBJECTS: [""],
            SELECTED_BEHAVIORS: ["p", "s"],
            config.INCLUDE_MODIFIERS: False,
            config.EXCLUDE_BEHAVIORS: False,
            config.START_TIME: 0,
            config.END_TIME: 100.0,
        }

        behav_seq_separator = "|"

        r, msg = export_observation.observation_to_behavioral_sequences(
            pj,
            selected_observations=observations,
            parameters=parameters,
            behaviors_separator=behav_seq_separator,
            separated_subjects=True,
            timed=False,
            file_name="output/Test_events_to_behavioral_sequences_test_6_multi_rows_description.txt",
        )

        assert (
            open("output/Test_events_to_behavioral_sequences_test_6_multi_rows_description.txt").read()
            == open("files/Test_events_to_behavioral_sequences_test_6_multi_rows_description.txt").read()
        )

    @pytest.mark.usefixtures("before")
    def test_7_all_subjects_2_observations(self):
        """
        all subjects
        2 observations
        2 behaviors
        """

        pj = json.loads(open("files/test.boris").read())

        observations = ["live export behavioral sequences", "observation #1"]
        parameters = {
            SELECTED_SUBJECTS: ["", "subject1", "subject2"],
            SELECTED_BEHAVIORS: ["p", "s"],
            config.INCLUDE_MODIFIERS: False,
            config.EXCLUDE_BEHAVIORS: False,
            config.START_TIME: 0,
            config.END_TIME: 100.0,
        }

        behav_seq_separator = "|"

        r, msg = export_observation.observation_to_behavioral_sequences(
            pj,
            selected_observations=observations,
            parameters=parameters,
            behaviors_separator=behav_seq_separator,
            separated_subjects=True,
            timed=False,
            file_name="output/Test_events_to_behavioral_sequences_test_7.txt",
        )

        assert (
            open("output/Test_events_to_behavioral_sequences_test_7.txt").read()
            == open("files/Test_events_to_behavioral_sequences_test_7.txt").read()
        )


"""
open("1", "w").write(out)

a= Test_export_events()
a.test_export_tabular_xlsx()
"""
